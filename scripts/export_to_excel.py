#!/usr/bin/env python3
"""
Test the Eureka Network scraper and export results to Excel for review.
"""

import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.scraper import EurekaNetworkScraper


def export_to_excel(grants: list, output_path: str = "scraper_results.xlsx"):
    """Export scraped grants to Excel for review."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eureka Grants"

    # Define headers
    headers = [
        "ID",
        "Title",
        "URL",
        "Status",
        "Programme",
        "Open Date",
        "Close Date",
        "Is Supplemental",
        "Funding Info",
        "Description (truncated)",
    ]

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data
    for row, grant in enumerate(grants, 2):
        ws.cell(row=row, column=1, value=grant.get('id', ''))
        ws.cell(row=row, column=2, value=grant.get('title', ''))
        ws.cell(row=row, column=3, value=grant.get('url', ''))
        ws.cell(row=row, column=4, value=grant.get('status', ''))
        ws.cell(row=row, column=5, value=grant.get('programme', ''))
        ws.cell(row=row, column=6, value=grant.get('open_date', ''))
        ws.cell(row=row, column=7, value=grant.get('close_date', ''))
        ws.cell(row=row, column=8, value='Yes' if grant.get('is_supplemental') else 'No')

        # Get funding info from raw
        raw = grant.get('raw', {})
        ws.cell(row=row, column=9, value=raw.get('funding_info', ''))

        # Truncate description for readability
        description = raw.get('description', '')
        if len(description) > 500:
            description = description[:500] + '...'
        ws.cell(row=row, column=10, value=description)

    # Auto-adjust column widths
    column_widths = [30, 50, 60, 10, 20, 15, 15, 15, 30, 80]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save
    wb.save(output_path)
    print(f"Saved to {output_path}")
    return output_path


def main():
    print("=" * 60)
    print("EUREKA NETWORK SCRAPER TEST")
    print("=" * 60)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Run scraper
    scraper = EurekaNetworkScraper()
    grants = scraper.scrape_all()

    print()
    print("=" * 60)
    print("RESULTS SUMMARY")
    print("=" * 60)
    print(f"Total grants scraped: {len(grants)}")

    # Count by status
    status_counts = {}
    for g in grants:
        status = g.get('status', 'Unknown')
        status_counts[status] = status_counts.get(status, 0) + 1

    print("\nBy status:")
    for status, count in sorted(status_counts.items()):
        print(f"  - {status}: {count}")

    # Count supplemental vs primary
    supplemental = sum(1 for g in grants if g.get('is_supplemental'))
    primary = len(grants) - supplemental
    print(f"\nBy type:")
    print(f"  - Primary R&D: {primary}")
    print(f"  - Supplemental: {supplemental}")

    # Export to Excel
    print()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Create output directory
    output_dir = Path(__file__).parent.parent / "outputs" / "excel"
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_path = output_dir / f"scraper_results_{timestamp}.xlsx"
    export_to_excel(grants, str(excel_path))

    # Also save JSON for reference
    json_path = output_dir / f"scraper_results_{timestamp}.json"
    with open(json_path, 'w') as f:
        json.dump(grants, f, indent=2)
    print(f"Saved JSON to {json_path}")

    print()
    print(f"Finished at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
